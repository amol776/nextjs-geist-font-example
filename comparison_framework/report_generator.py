import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
import openpyxl
from openpyxl.styles import PatternFill, Font
from utils import log_error, format_timestamp

class ReportGenerator:
    """Class to handle generation of comparison reports"""

    # Define colors for Excel reports
    PASS_COLOR = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
    FAIL_COLOR = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light pink
    HEADER_COLOR = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')  # Steel blue
    HEADER_FONT = Font(color='FFFFFF', bold=True)  # White, bold

    @staticmethod
    def generate_diff_report(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        column_mapping: Dict[str, str],
        exclude_columns: List[str] = None,
        join_keys: Optional[List[str]] = None
    ) -> Tuple[pd.DataFrame, bool]:
        """
        Generate a difference report between source and target DataFrames.
        
        Args:
            source_df: Source DataFrame
            target_df: Target DataFrame
            column_mapping: Dictionary mapping source columns to target columns
            exclude_columns: List of columns to exclude from comparison
            join_keys: Optional list of columns to use as join keys
            
        Returns:
            Tuple[pd.DataFrame, bool]: (Difference report DataFrame, Whether differences were found)
        """
        try:
            # Filter out excluded columns
            if exclude_columns:
                column_mapping = {k: v for k, v in column_mapping.items() 
                                if k not in exclude_columns}

            # Convert data types for comparison
            source_df = source_df.copy()
            target_df = target_df.copy()
            
            # Function to safely convert numeric columns
            def safe_numeric_conversion(series):
                try:
                    if series.dtype == 'object':
                        return pd.to_numeric(series, errors='coerce')
                    return series
                except:
                    return series

            # Apply numeric conversion to both dataframes
            for source_col, target_col in column_mapping.items():
                source_df[source_col] = safe_numeric_conversion(source_df[source_col])
                target_df[target_col] = safe_numeric_conversion(target_df[target_col])

            differences = []
            
            if join_keys and all(key in source_df.columns and key in target_df.columns for key in join_keys):
                # Merge source and target on join keys
                merged_df = pd.merge(
                    source_df,
                    target_df,
                    how='outer',
                    left_on=join_keys,
                    right_on=join_keys,
                    indicator=True
                )
                
                # Compare mapped columns for merged data
                for source_col, target_col in column_mapping.items():
                    source_values = merged_df[source_col + '_x']
                    target_values = merged_df[target_col + '_y']
                    
                    # Handle missing values from outer join
                    for idx in merged_df.index:
                        if merged_df.at[idx, '_merge'] == 'left_only':
                            differences.append({
                                'Column': source_col,
                                'Join_Keys': {k: merged_df.at[idx, k] for k in join_keys},
                                'Source_Value': source_values[idx],
                                'Target_Value': 'Missing in Target'
                            })
                        elif merged_df.at[idx, '_merge'] == 'right_only':
                            differences.append({
                                'Column': source_col,
                                'Join_Keys': {k: merged_df.at[idx, k] for k in join_keys},
                                'Source_Value': 'Missing in Source',
                                'Target_Value': target_values[idx]
                            })
                        elif pd.notna(source_values[idx]) and pd.notna(target_values[idx]):
                            if source_values[idx] != target_values[idx]:
                                differences.append({
                                    'Column': source_col,
                                    'Join_Keys': {k: merged_df.at[idx, k] for k in join_keys},
                                    'Source_Value': source_values[idx],
                                    'Target_Value': target_values[idx]
                                })
            else:
                # Compare mapped columns without join keys
                for source_col, target_col in column_mapping.items():
                    source_values = source_df[source_col]
                    target_values = target_df[target_col]
                    
                    # Find rows where values don't match
                    for idx in range(min(len(source_values), len(target_values))):
                        if pd.notna(source_values.iloc[idx]) and pd.notna(target_values.iloc[idx]):
                            if source_values.iloc[idx] != target_values.iloc[idx]:
                                differences.append({
                                    'Column': source_col,
                                    'Row_Index': idx,
                                    'Source_Value': source_values.iloc[idx],
                                    'Target_Value': target_values.iloc[idx]
                                })

            # Create difference report DataFrame
            if differences:
                diff_df = pd.DataFrame(differences)
                return diff_df, True
            else:
                return pd.DataFrame({'Message': ['No differences found']}, index=[0]), False

        except Exception as e:
            log_error(f"Error generating difference report: {str(e)}")
            return pd.DataFrame(), False

    @staticmethod
    def generate_profiling_report(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        column_mapping: Dict[str, str],
        join_keys: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        Generate Y-Data profiling comparison report.
        
        Args:
            source_df: Source DataFrame
            target_df: Target DataFrame
            column_mapping: Dictionary mapping source columns to target columns
            join_keys: Optional list of columns to use as join keys
            
        Returns:
            pd.DataFrame: Profiling report DataFrame
        """
        try:
            # Convert data types based on mapping
            source_df = source_df.copy()
            target_df = target_df.copy()
            
            # Function to safely convert numeric columns
            def safe_numeric_conversion(series):
                try:
                    # Try to convert to numeric, coercing errors to NaN
                    return pd.to_numeric(series, errors='coerce')
                except:
                    return series

            # Apply numeric conversion to both dataframes
            for source_col, target_col in column_mapping.items():
                # Convert source column if it looks numeric
                if source_df[source_col].dtype == 'object':
                    source_df[source_col] = safe_numeric_conversion(source_df[source_col])
                
                # Convert target column if it looks numeric
                if target_df[target_col].dtype == 'object':
                    target_df[target_col] = safe_numeric_conversion(target_df[target_col])

            # If join keys are provided, merge the dataframes
            if join_keys and all(key in source_df.columns and key in target_df.columns for key in join_keys):
                merged_df = pd.merge(
                    source_df,
                    target_df,
                    how='inner',
                    left_on=join_keys,
                    right_on=join_keys,
                    suffixes=('_source', '_target')
                )
            else:
                # If no join keys or invalid keys, use the original dataframes
                merged_df = None

            profile_data = []
            
            for source_col, target_col in column_mapping.items():
                # Get source and target series
                source_series = source_df[source_col]
                target_series = target_df[target_col]

                # Calculate basic statistics
                source_stats = {
                    'count': len(source_series),
                    'null_count': source_series.isnull().sum(),
                    'distinct_count': source_series.nunique()
                }
                
                target_stats = {
                    'count': len(target_series),
                    'null_count': target_series.isnull().sum(),
                    'distinct_count': target_series.nunique()
                }

                # Calculate numeric statistics if possible
                for series, stats, prefix in [(source_series, source_stats, 'source'), 
                                           (target_series, target_stats, 'target')]:
                    try:
                        numeric_series = pd.to_numeric(series, errors='coerce')
                        stats.update({
                            'min': numeric_series.min(),
                            'max': numeric_series.max(),
                            'mean': numeric_series.mean(),
                            'std': numeric_series.std(),
                            'median': numeric_series.median()
                        })
                    except:
                        stats.update({
                            'min': 'N/A',
                            'max': 'N/A',
                            'mean': 'N/A',
                            'std': 'N/A',
                            'median': 'N/A'
                        })

                # Calculate match percentage
                if merged_df is not None:
                    source_col_merged = f"{source_col}_source"
                    target_col_merged = f"{target_col}_target"
                    matching_values = (merged_df[source_col_merged] == merged_df[target_col_merged]).sum()
                    total_rows = len(merged_df)
                    match_percentage = (matching_values / total_rows * 100) if total_rows > 0 else 0
                else:
                    # If no merge possible, calculate based on position
                    min_len = min(len(source_series), len(target_series))
                    if min_len > 0:
                        matching_values = (source_series.iloc[:min_len] == target_series.iloc[:min_len]).sum()
                        match_percentage = (matching_values / min_len * 100)
                    else:
                        match_percentage = 0

                # Format numeric values as strings with appropriate precision
                def format_value(value):
                    if pd.isna(value) or value == 'N/A':
                        return 'N/A'
                    try:
                        if isinstance(value, (int, float)):
                            return f"{value:,.2f}" if value % 1 != 0 else f"{int(value):,}"
                        return str(value)
                    except:
                        return str(value)

                profile_data.append({
                    'Column': source_col,
                    'Source_Count': format_value(source_stats['count']),
                    'Target_Count': format_value(target_stats['count']),
                    'Source_Nulls': format_value(source_stats['null_count']),
                    'Target_Nulls': format_value(target_stats['null_count']),
                    'Source_Distinct': format_value(source_stats['distinct_count']),
                    'Target_Distinct': format_value(target_stats['distinct_count']),
                    'Source_Min': format_value(source_stats['min']),
                    'Target_Min': format_value(target_stats['min']),
                    'Source_Max': format_value(source_stats['max']),
                    'Target_Max': format_value(target_stats['max']),
                    'Source_Mean': format_value(source_stats['mean']),
                    'Target_Mean': format_value(target_stats['mean']),
                    'Source_StdDev': format_value(source_stats['std']),
                    'Target_StdDev': format_value(target_stats['std']),
                    'Source_Median': format_value(source_stats['median']),
                    'Target_Median': format_value(target_stats['median']),
                    'Match_Percentage': f"{match_percentage:.2f}%"
                })
            
            # Create DataFrame and sort columns for better readability
            df = pd.DataFrame(profile_data)
            column_order = [
                'Column',
                'Source_Count', 'Target_Count',
                'Source_Nulls', 'Target_Nulls',
                'Source_Distinct', 'Target_Distinct',
                'Source_Min', 'Target_Min',
                'Source_Max', 'Target_Max',
                'Source_Mean', 'Target_Mean',
                'Source_Median', 'Target_Median',
                'Source_StdDev', 'Target_StdDev',
                'Match_Percentage'
            ]
            return df[column_order]

        except Exception as e:
            log_error(f"Error generating profiling report: {str(e)}")
            return pd.DataFrame()

    @staticmethod
    def _calculate_column_stats(series: pd.Series) -> Dict[str, Any]:
        """Calculate statistical measures for a column."""
        stats = {
            'count': len(series),
            'null_count': series.isnull().sum(),
            'distinct_count': series.nunique(),
            'min': series.min() if series.dtype in ['int64', 'float64'] else str(series.min()),
            'max': series.max() if series.dtype in ['int64', 'float64'] else str(series.max())
        }
        
        if series.dtype in ['int64', 'float64']:
            stats['mean'] = series.mean()
            
        return stats

    @staticmethod
    def generate_regression_report(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        column_mapping: Dict[str, str],
        output_path: str
    ) -> bool:
        """
        Generate Excel-based regression report with multiple tabs.
        
        Args:
            source_df: Source DataFrame
            target_df: Target DataFrame
            column_mapping: Dictionary mapping source columns to target columns
            output_path: Path to save the Excel report
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Generate AggregationCheck tab
                ReportGenerator._generate_aggregation_check(
                    source_df, target_df, column_mapping, writer)
                
                # Generate CountCheck tab
                ReportGenerator._generate_count_check(
                    source_df, target_df, writer)
                
                # Generate DistinctCheck tab
                ReportGenerator._generate_distinct_check(
                    source_df, target_df, column_mapping, writer)
                
                # Apply formatting
                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    ReportGenerator._apply_excel_formatting(workbook[sheet_name])
                
            return True

        except Exception as e:
            log_error(f"Error generating regression report: {str(e)}")
            return False

    @staticmethod
    def _generate_aggregation_check(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        column_mapping: Dict[str, str],
        writer: pd.ExcelWriter
    ) -> None:
        """Generate AggregationCheck tab in the regression report."""
        agg_data = []
        
        for source_col, target_col in column_mapping.items():
            if source_df[source_col].dtype in ['int64', 'float64']:
                source_sum = source_df[source_col].sum()
                target_sum = target_df[target_col].sum()
                
                agg_data.append({
                    'Source_Column': source_col,
                    'Target_Column': target_col,
                    'Source_Sum': source_sum,
                    'Target_Sum': target_sum,
                    'Result': 'PASS' if np.isclose(source_sum, target_sum) else 'FAIL'
                })
        
        pd.DataFrame(agg_data).to_excel(writer, sheet_name='AggregationCheck', index=False)

    @staticmethod
    def _generate_count_check(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        writer: pd.ExcelWriter
    ) -> None:
        """Generate CountCheck tab in the regression report."""
        count_data = [{
            'Source_File_Name': 'Source Dataset',
            'Source_Path': 'N/A',
            'Data_Count': len(source_df),
            'Target_Count': len(target_df),
            'Result': 'PASS' if len(source_df) == len(target_df) else 'FAIL'
        }]
        
        pd.DataFrame(count_data).to_excel(writer, sheet_name='CountCheck', index=False)

    @staticmethod
    def _generate_distinct_check(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        column_mapping: Dict[str, str],
        writer: pd.ExcelWriter
    ) -> None:
        """Generate DistinctCheck tab in the regression report."""
        distinct_data = []
        
        for source_col, target_col in column_mapping.items():
            if source_df[source_col].dtype not in ['int64', 'float64']:
                source_distinct = set(source_df[source_col].dropna().unique())
                target_distinct = set(target_df[target_col].dropna().unique())
                
                distinct_data.append({
                    'Column': source_col,
                    'Source_Distinct_Count': len(source_distinct),
                    'Target_Distinct_Count': len(target_distinct),
                    'Source_Distinct_Values': ', '.join(map(str, sorted(source_distinct))),
                    'Target_Distinct_Values': ', '.join(map(str, sorted(target_distinct))),
                    'Result': 'PASS' if source_distinct == target_distinct else 'FAIL'
                })
        
        pd.DataFrame(distinct_data).to_excel(writer, sheet_name='DistinctCheck', index=False)

    @staticmethod
    def _apply_excel_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Apply formatting to Excel worksheet."""
        # Format headers
        for cell in worksheet[1]:
            cell.fill = ReportGenerator.HEADER_COLOR
            cell.font = ReportGenerator.HEADER_FONT

        # Format PASS/FAIL cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value == 'PASS':
                    cell.fill = ReportGenerator.PASS_COLOR
                elif cell.value == 'FAIL':
                    cell.fill = ReportGenerator.FAIL_COLOR

    @staticmethod
    def generate_side_by_side_report(
        source_df: pd.DataFrame,
        target_df: pd.DataFrame,
        column_mapping: Dict[str, str],
        output_path: Optional[str] = None,
        join_keys: Optional[List[str]] = None
    ) -> Tuple[pd.DataFrame, str]:
        """
        Generate side-by-side difference report.
        
        Args:
            source_df: Source DataFrame
            target_df: Target DataFrame
            column_mapping: Dictionary mapping source columns to target columns
            output_path: Optional path to save Excel report
            join_keys: Optional list of columns to use as join keys
            
        Returns:
            Tuple[pd.DataFrame, str]: (Difference report DataFrame, Output path if saved)
        """
        try:
            # Create side-by-side comparison
            if join_keys and all(key in source_df.columns and key in target_df.columns for key in join_keys):
                # Prepare source dataframe
                source_subset = source_df.copy()
                source_cols = list(column_mapping.keys())
                source_subset.columns = [f"{col}_source" if col in source_cols else col for col in source_subset.columns]
                
                # Prepare target dataframe
                target_subset = target_df.copy()
                target_cols = list(column_mapping.values())
                rename_dict = {v: f"{k}_target" for k, v in column_mapping.items()}
                target_subset.rename(columns=rename_dict, inplace=True)
                
                # Merge dataframes
                comparison_df = pd.merge(
                    source_subset,
                    target_subset,
                    on=join_keys,
                    how='outer',
                    indicator=True,
                    suffixes=('_source', '_target')
                )
                
                # Add match status column
                comparison_df['Match_Status'] = comparison_df.apply(
                    lambda row: 'Match' if row['_merge'] == 'both' else 'Source Only' if row['_merge'] == 'left_only' else 'Target Only',
                    axis=1
                )
                
                # Drop the merge indicator column
                comparison_df.drop('_merge', axis=1, inplace=True)
                
            else:
                # Without join keys, do a simple side-by-side comparison
                max_rows = min(len(source_df), len(target_df))
                comparison_data = []
                
                for idx in range(max_rows):
                    row_data = {'Row_Index': idx}
                    has_difference = False
                    
                    for source_col, target_col in column_mapping.items():
                        source_val = source_df[source_col].iloc[idx]
                        target_val = target_df[target_col].iloc[idx]
                        
                        row_data[f'Source_{source_col}'] = source_val
                        row_data[f'Target_{target_col}'] = target_val
                        
                        if pd.notna(source_val) and pd.notna(target_val):
                            if source_val != target_val:
                                has_difference = True
                    
                    if has_difference:
                        comparison_data.append(row_data)

                if not comparison_data:
                    comparison_df = pd.DataFrame({'Message': ['No differences found']})
                else:
                    comparison_df = pd.DataFrame(comparison_data)

            # Save to Excel if path provided
            if output_path:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    comparison_df.to_excel(writer, index=False)
                    ReportGenerator._apply_excel_formatting(writer.book.active)
                return comparison_df, output_path
            
            return comparison_df, ''

        except Exception as e:
            log_error(f"Error generating side-by-side report: {str(e)}")
            return pd.DataFrame(), ''
